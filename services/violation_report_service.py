"""
Violation Report Service

Generates tabular reports of SOD violations with export options
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
import json

logger = logging.getLogger(__name__)


class ViolationReportService:
    """Service for generating violation reports in various formats"""

    def __init__(self):
        """Initialize the report service"""
        pass

    def generate_markdown_table(
        self,
        violations: List[Dict[str, Any]],
        limit: int = 5
    ) -> str:
        """
        Generate a markdown table of violations

        Args:
            violations: List of violation dictionaries
            limit: Maximum number of violations to show

        Returns:
            Markdown formatted table string
        """
        if not violations:
            return "No violations found."

        # Sort by severity (CRITICAL > HIGH > MEDIUM > LOW)
        severity_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
        sorted_violations = sorted(
            violations,
            key=lambda v: severity_order.get(v.get('severity', 'LOW'), 4)
        )

        # Take top N violations
        top_violations = sorted_violations[:limit]

        # Build markdown table
        output = f"**Top {len(top_violations)} Violations (of {len(violations)} total):**\n\n"
        output += "| # | Violation Type | Severity | Conflicting Roles | Description |\n"
        output += "|---|----------------|----------|-------------------|-------------|\n"

        for i, violation in enumerate(top_violations, 1):
            title = violation.get('title', 'Unknown')
            severity = violation.get('severity', 'UNKNOWN')

            # Get conflicting roles
            roles = violation.get('conflicting_roles', [])
            if isinstance(roles, str):
                try:
                    roles = json.loads(roles)
                except:
                    roles = [roles]

            roles_str = f"{len(roles)} roles"
            if len(roles) <= 2:
                roles_str = ", ".join(roles)

            # Get description (truncate if too long)
            desc = violation.get('description', '')
            if len(desc) > 80:
                desc = desc[:77] + "..."

            # Severity emoji
            severity_emoji = {
                "CRITICAL": "🔴",
                "HIGH": "🟠",
                "MEDIUM": "🟡",
                "LOW": "🟢"
            }.get(severity, "⚪")

            output += f"| {i} | {title} | {severity_emoji} {severity} | {roles_str} | {desc} |\n"

        if len(violations) > limit:
            output += f"\n_Showing {limit} of {len(violations)} violations. Use export to see all._\n"

        return output

    def generate_detailed_table(
        self,
        violations: List[Dict[str, Any]],
        limit: int = 10
    ) -> str:
        """
        Generate a detailed text table with role matrix

        Args:
            violations: List of violation dictionaries
            limit: Maximum number of violations to show

        Returns:
            Formatted text table with role matrix
        """
        if not violations:
            return "No violations found."

        # Sort by severity
        severity_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
        sorted_violations = sorted(
            violations,
            key=lambda v: severity_order.get(v.get('severity', 'LOW'), 4)
        )

        # Get all unique roles across all violations
        all_roles = set()
        for v in sorted_violations[:limit]:
            roles = v.get('conflicting_roles', [])
            if isinstance(roles, str):
                try:
                    roles = json.loads(roles)
                except:
                    roles = [roles]
            all_roles.update(roles)

        all_roles = sorted(list(all_roles))

        # Build detailed table
        output = f"**Violation-Role Matrix (Top {limit} of {len(violations)}):**\n\n"

        # Header
        output += "| Violation Type | Severity |"
        for role in all_roles:
            role_short = role[:15] + "..." if len(role) > 15 else role
            output += f" {role_short} |"
        output += "\n"

        # Separator
        output += "|" + "|".join(["---"] * (2 + len(all_roles))) + "|\n"

        # Rows
        for violation in sorted_violations[:limit]:
            title = violation.get('title', 'Unknown')
            # Truncate long titles
            if len(title) > 40:
                title = title[:37] + "..."

            severity = violation.get('severity', 'UNKNOWN')
            severity_emoji = {
                "CRITICAL": "🔴",
                "HIGH": "🟠",
                "MEDIUM": "🟡",
                "LOW": "🟢"
            }.get(severity, "⚪")

            output += f"| {title} | {severity_emoji} {severity} |"

            # Check if each role is involved in this violation
            v_roles = violation.get('conflicting_roles', [])
            if isinstance(v_roles, str):
                try:
                    v_roles = json.loads(v_roles)
                except:
                    v_roles = [v_roles]

            for role in all_roles:
                if role in v_roles:
                    output += " ✓ |"
                else:
                    output += " |"
            output += "\n"

        return output

    def export_to_csv(
        self,
        violations: List[Dict[str, Any]],
        output_path: str
    ) -> str:
        """
        Export violations to CSV file

        Args:
            violations: List of violation dictionaries
            output_path: Path to save CSV file

        Returns:
            Success message with file path
        """
        import csv

        if not violations:
            return "No violations to export."

        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = [
                    'Violation Type',
                    'Severity',
                    'Risk Score',
                    'Status',
                    'Conflicting Roles',
                    'Conflicting Permissions',
                    'Description',
                    'Detected Date'
                ]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                writer.writeheader()
                for v in violations:
                    # Parse JSON fields
                    roles = v.get('conflicting_roles', [])
                    if isinstance(roles, str):
                        try:
                            roles = json.loads(roles)
                        except:
                            pass

                    perms = v.get('conflicting_permissions', [])
                    if isinstance(perms, str):
                        try:
                            perms = json.loads(perms)
                        except:
                            pass

                    writer.writerow({
                        'Violation Type': v.get('title', ''),
                        'Severity': v.get('severity', ''),
                        'Risk Score': v.get('risk_score', 0),
                        'Status': v.get('status', ''),
                        'Conflicting Roles': ', '.join(roles) if isinstance(roles, list) else str(roles),
                        'Conflicting Permissions': ', '.join(perms) if isinstance(perms, list) else str(perms),
                        'Description': v.get('description', ''),
                        'Detected Date': v.get('detected_at', '')
                    })

            return f"✅ Exported {len(violations)} violations to: {output_path}"

        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return f"❌ Error exporting to CSV: {str(e)}"

    def export_to_excel(
        self,
        violations: List[Dict[str, Any]],
        output_path: str,
        user_name: Optional[str] = None
    ) -> str:
        """
        Export violations to Excel file with formatting

        Args:
            violations: List of violation dictionaries
            output_path: Path to save Excel file
            user_name: Optional user name for report title

        Returns:
            Success message with file path
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return "❌ Excel export requires: pip install pandas openpyxl"

        if not violations:
            return "No violations to export."

        try:
            # Prepare data for DataFrame
            data = []
            for v in violations:
                # Parse JSON fields
                roles = v.get('conflicting_roles', [])
                if isinstance(roles, str):
                    try:
                        roles = json.loads(roles)
                    except:
                        pass

                perms = v.get('conflicting_permissions', [])
                if isinstance(perms, str):
                    try:
                        perms = json.loads(perms)
                    except:
                        pass

                data.append({
                    'Violation Type': v.get('title', ''),
                    'Severity': v.get('severity', ''),
                    'Risk Score': v.get('risk_score', 0),
                    'Status': v.get('status', ''),
                    'Conflicting Roles': ', '.join(roles) if isinstance(roles, list) else str(roles),
                    'Role Count': len(roles) if isinstance(roles, list) else 0,
                    'Conflicting Permissions': ', '.join(perms) if isinstance(perms, list) else str(perms),
                    'Description': v.get('description', ''),
                    'Detected Date': v.get('detected_at', '')
                })

            # Create DataFrame
            df = pd.DataFrame(data)

            # Write to Excel
            df.to_excel(output_path, index=False, sheet_name='Violations')

            # Apply formatting
            wb = load_workbook(output_path)
            ws = wb.active

            # Format header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Color code severity
            severity_colors = {
                "CRITICAL": "FFC7CE",  # Red
                "HIGH": "FFD966",      # Orange
                "MEDIUM": "FFEB9C",    # Yellow
                "LOW": "C6EFCE"        # Green
            }

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                for cell in row:
                    severity = cell.value
                    if severity in severity_colors:
                        cell.fill = PatternFill(
                            start_color=severity_colors[severity],
                            end_color=severity_colors[severity],
                            fill_type="solid"
                        )

            # Adjust column widths
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 50
            ws.column_dimensions['H'].width = 60
            ws.column_dimensions['I'].width = 20

            # Add metadata sheet
            ws_meta = wb.create_sheet("Report Info")
            ws_meta['A1'] = "SOD Violation Report"
            ws_meta['A1'].font = Font(bold=True, size=14)
            ws_meta['A3'] = "Generated:"
            ws_meta['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if user_name:
                ws_meta['A4'] = "User:"
                ws_meta['B4'] = user_name
            ws_meta['A5'] = "Total Violations:"
            ws_meta['B5'] = len(violations)
            ws_meta['A7'] = "Severity Breakdown:"

            # Count by severity
            severity_counts = {}
            for v in violations:
                sev = v.get('severity', 'UNKNOWN')
                severity_counts[sev] = severity_counts.get(sev, 0) + 1

            row = 8
            for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
                if severity in severity_counts:
                    ws_meta[f'A{row}'] = f"  {severity}:"
                    ws_meta[f'B{row}'] = severity_counts[severity]
                    row += 1

            wb.save(output_path)

            summary = f"✅ Exported {len(violations)} violations to Excel: {output_path}\n"
            summary += f"   • CRITICAL: {severity_counts.get('CRITICAL', 0)}\n"
            summary += f"   • HIGH: {severity_counts.get('HIGH', 0)}\n"
            summary += f"   • MEDIUM: {severity_counts.get('MEDIUM', 0)}\n"
            summary += f"   • LOW: {severity_counts.get('LOW', 0)}"

            return summary

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}", exc_info=True)
            return f"❌ Error exporting to Excel: {str(e)}"

    def export_to_google_sheets(
        self,
        violations: List[Dict[str, Any]],
        sheet_name: str,
        user_name: Optional[str] = None
    ) -> str:
        """
        Export violations to Google Sheets

        Args:
            violations: List of violation dictionaries
            sheet_name: Name for the Google Sheet
            user_name: Optional user name for report title

        Returns:
            Success message with sheet URL
        """
        try:
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
        except ImportError:
            return "❌ Google Sheets export requires: pip install gspread oauth2client"

        # TODO: Implement Google Sheets export
        # Requires service account credentials setup
        return "⚠️ Google Sheets export not yet implemented. Use Excel export instead."
